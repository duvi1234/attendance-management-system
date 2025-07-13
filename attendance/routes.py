from flask import Blueprint, render_template, redirect, url_for, request, session, flash, send_file
from models.models import db, Advisor, Batch, Student, Subject, Semester, Attendance
from sqlalchemy.sql import func
import pandas as pd
import io

attendance_bp = Blueprint('attendance', __name__, url_prefix='/attendance')

# Ensure advisor is logged in
def login_required(view):
    def wrapped_view(**kwargs):
        if 'advisor_id' not in session:
            return redirect(url_for('auth.login'))
        return view(**kwargs)
    wrapped_view.__name__ = view.__name__
    return wrapped_view

# Dashboard – shows students in advisor's batch with semester-wise attendance
@attendance_bp.route('/dashboard')
@login_required
def dashboard():
    advisor = Advisor.query.get(session['advisor_id'])
    students = Student.query.filter_by(batch_id=advisor.batch_id).all()
    semesters = Semester.query.filter_by(batch_id=advisor.batch_id).all()
    active_semester = Semester.query.filter_by(batch_id=advisor.batch_id, is_active=True).first()
    
    # Get selected semester (default to active semester)
    selected_semester_id = request.args.get('semester_id', type=int)
    if selected_semester_id:
        selected_semester = Semester.query.get(selected_semester_id)
        if not selected_semester or selected_semester.batch_id != advisor.batch_id:
            selected_semester = active_semester
    else:
        selected_semester = active_semester
    
    # Get view type (overall or subject-wise)
    view_type = request.args.get('view_type', 'overall')
    subject_id = request.args.get('subject_id', type=int)
    selected_subject = None
    if subject_id:
        selected_subject = Subject.query.get(subject_id)

    student_data = []
    subjects_data = []
    
    if selected_semester:
        # Get subjects for the selected semester
        subjects = Subject.query.filter_by(semester_id=selected_semester.id).all()
        
        for student in students:
            if view_type == 'subject_wise':
                # Subject-wise view
                student_subjects = []
                total_attendance = 0
                total_classes = 0
                filtered_subjects = subjects
                if subject_id:
                    filtered_subjects = [s for s in subjects if s.id == subject_id]
                for subject in filtered_subjects:
                    attendance_records = Attendance.query.filter_by(
                        student_id=student.id, 
                        subject_id=subject.id
                    ).all()
                    subject_total_classes = sum(a.total_classes for a in attendance_records)
                    subject_attended_classes = sum(a.attended_classes for a in attendance_records)
                    subject_percentage = round((subject_attended_classes / subject_total_classes) * 100, 2) if subject_total_classes > 0 else 0
                    student_subjects.append({
                        'name': subject.name,
                        'total_classes': subject_total_classes,
                        'attended_classes': subject_attended_classes,
                        'percentage': subject_percentage
                    })
                    total_classes += subject_total_classes
                    total_attendance += subject_attended_classes
                overall_percentage = round((total_attendance / total_classes) * 100, 2) if total_classes > 0 else 0
                student_data.append({
                    'id': student.id,
                    'name': student.name,
                    'roll_no': student.roll_no,
                    'reg_no': student.reg_no,
                    'subjects': student_subjects,
                    'total_classes': total_classes,
                    'total_attended': total_attendance,
                    'overall_percentage': overall_percentage
                })
            else:
                # Overall view (all subjects combined)
                semester_subjects = Subject.query.filter_by(semester_id=selected_semester.id).all()
                semester_total_classes = 0
                semester_attended_classes = 0
                
                for subject in semester_subjects:
                    attendance_records = Attendance.query.filter_by(
                        student_id=student.id, 
                        subject_id=subject.id
                    ).all()
                    
                    subject_total_classes = sum(a.total_classes for a in attendance_records)
                    subject_attended_classes = sum(a.attended_classes for a in attendance_records)
                    
                    semester_total_classes += subject_total_classes
                    semester_attended_classes += subject_attended_classes
                
                semester_percentage = round((semester_attended_classes / semester_total_classes) * 100, 2) if semester_total_classes > 0 else 0
                
                student_data.append({
                    'id': student.id,
                    'name': student.name,
                    'roll_no': student.roll_no,
                    'reg_no': student.reg_no,
                    'total_classes': semester_total_classes,
                    'total_attended': semester_attended_classes,
                    'percentage': semester_percentage
                })
    else:
        # Show overall data for all semesters (original behavior)
        for student in students:
            student_semesters = []
            total_attendance = 0
            total_classes = 0
            
            for semester in semesters:
                semester_subjects = Subject.query.filter_by(semester_id=semester.id).all()
                semester_total_classes = 0
                semester_attended_classes = 0
                
                for subject in semester_subjects:
                    attendance_records = Attendance.query.filter_by(
                        student_id=student.id, 
                        subject_id=subject.id
                    ).all()
                    
                    subject_total_classes = sum(a.total_classes for a in attendance_records)
                    subject_attended_classes = sum(a.attended_classes for a in attendance_records)
                    
                    semester_total_classes += subject_total_classes
                    semester_attended_classes += subject_attended_classes
                
                semester_percentage = round((semester_attended_classes / semester_total_classes) * 100, 2) if semester_total_classes > 0 else 0
                
                student_semesters.append({
                    'name': semester.name,
                    'total_classes': semester_total_classes,
                    'attended_classes': semester_attended_classes,
                    'percentage': semester_percentage,
                    'is_active': semester.is_active
                })
                
                total_classes += semester_total_classes
                total_attendance += semester_attended_classes
            
            overall_percentage = round((total_attendance / total_classes) * 100, 2) if total_classes > 0 else 0
            
            student_data.append({
                'id': student.id,
                'name': student.name,
                'roll_no': student.roll_no,
                'reg_no': student.reg_no,
                'semesters': student_semesters,
                'total_classes': total_classes,
                'total_attended': total_attendance,
                'overall_percentage': overall_percentage
            })

    return render_template('dashboard.html', 
                         students=student_data, 
                         semesters=semesters, 
                         active_semester=active_semester,
                         selected_semester=selected_semester,
                         view_type=view_type,
                         subject_id=subject_id,
                         selected_subject=selected_subject
                         )

# Manage semesters
@attendance_bp.route('/semesters', methods=['GET', 'POST'])
@login_required
def manage_semesters():
    advisor = Advisor.query.get(session['advisor_id'])
    semesters = Semester.query.filter_by(batch_id=advisor.batch_id).all()
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add_semester':
            semester_name = request.form.get('semester_name')
            if semester_name:
                # Check if semester already exists
                existing_semester = Semester.query.filter_by(
                    name=semester_name, 
                    batch_id=advisor.batch_id
                ).first()
                
                if existing_semester:
                    flash('Semester already exists!', 'danger')
                else:
                    # If this is the first semester, make it active
                    is_active = len(semesters) == 0
                    semester = Semester(name=semester_name, batch_id=advisor.batch_id, is_active=is_active)
                    db.session.add(semester)
                    db.session.commit()
                    flash('Semester added successfully!', 'success')
        
        elif action == 'activate_semester':
            semester_id = request.form.get('semester_id')
            semester = Semester.query.get(semester_id)
            if semester and semester.batch_id == advisor.batch_id:
                # Deactivate all other semesters
                Semester.query.filter_by(batch_id=advisor.batch_id).update({'is_active': False})
                # Activate this semester
                semester.is_active = True
                db.session.commit()
                flash('Semester activated successfully!', 'success')
            else:
                flash('Invalid semester!', 'danger')
        
        elif action == 'delete_semester':
            semester_id = request.form.get('semester_id')
            semester = Semester.query.get(semester_id)
            if semester and semester.batch_id == advisor.batch_id:
                # Delete all subjects and attendance records for this semester
                subjects = Subject.query.filter_by(semester_id=semester_id).all()
                for subject in subjects:
                    Attendance.query.filter_by(subject_id=subject.id).delete()
                    db.session.delete(subject)
                db.session.delete(semester)
                db.session.commit()
                flash('Semester deleted successfully!', 'success')
            else:
                flash('Invalid semester!', 'danger')
    
    return render_template('manage_semesters.html', semesters=semesters)

# Manage subjects
@attendance_bp.route('/subjects', methods=['GET', 'POST'])
@login_required
def manage_subjects():
    advisor = Advisor.query.get(session['advisor_id'])
    active_semester = Semester.query.filter_by(batch_id=advisor.batch_id, is_active=True).first()
    
    if not active_semester:
        flash('Please create and activate a semester first!', 'warning')
        return redirect(url_for('attendance.manage_semesters'))
    
    subjects = Subject.query.filter_by(semester_id=active_semester.id).all()
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add_subject':
            subject_name = request.form.get('subject_name')
            if subject_name:
                # Check if subject already exists in this semester
                existing_subject = Subject.query.filter_by(
                    name=subject_name, 
                    semester_id=active_semester.id
                ).first()
                
                if existing_subject:
                    flash('Subject already exists in this semester!', 'danger')
                else:
                    subject = Subject(name=subject_name, semester_id=active_semester.id)
                    db.session.add(subject)
                    db.session.commit()
                    flash('Subject added successfully!', 'success')
        
        elif action == 'delete_subject':
            subject_id = request.form.get('subject_id')
            subject = Subject.query.get(subject_id)
            if subject and subject.semester_id == active_semester.id:
                # Delete all attendance records for this subject
                Attendance.query.filter_by(subject_id=subject_id).delete()
                db.session.delete(subject)
                db.session.commit()
                flash('Subject deleted successfully!', 'success')
            else:
                flash('Invalid subject!', 'danger')
    
    return render_template('manage_subjects.html', subjects=subjects, active_semester=active_semester)

# Form to add weekly attendance by subject
@attendance_bp.route('/update', methods=['GET', 'POST'])
@login_required
def update_attendance():
    advisor = Advisor.query.get(session['advisor_id'])
    students = Student.query.filter_by(batch_id=advisor.batch_id).all()
    active_semester = Semester.query.filter_by(batch_id=advisor.batch_id, is_active=True).first()
    
    if not active_semester:
        flash('Please create and activate a semester first!', 'warning')
        return redirect(url_for('attendance.manage_semesters'))
    
    subjects = Subject.query.filter_by(semester_id=active_semester.id).all()

    if request.method == 'POST':
        subject_id = int(request.form['subject'])
        week_date = request.form.get('week_date')
        
        # Get the week dates based on the selected date
        from datetime import datetime, timedelta
        if week_date:
            selected_date = datetime.strptime(week_date, '%Y-%m-%d')
            # Get Monday of the week
            days_since_monday = selected_date.weekday()
            monday = selected_date - timedelta(days=days_since_monday)
            
            # Get week dates (Monday to Friday, optionally Saturday)
            include_saturday = 'include_saturday' in request.form
            week_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
            if include_saturday:
                week_days.append('Saturday')
            
            # Process attendance for each student
            for student in students:
                total_key = f"total_{student.id}"
                attended_key = f"attended_{student.id}"
                
                total = int(request.form.get(total_key, 0))
                attended = int(request.form.get(attended_key, 0))
                
                if total > 0:  # Only save if there are classes
                    # Calculate classes per day
                    classes_per_day = total // len(week_days)
                    remainder = total % len(week_days)
                    
                    # Distribute classes across all days
                    for i, day in enumerate(week_days):
                        current_date = monday + timedelta(days=i)
                        date_str = current_date.strftime('%Y-%m-%d')
                        
                        # Check if this day is marked as a holiday
                        is_holiday = f"college_holiday_{day}" in request.form
                        
                        if not is_holiday:
                            # Calculate classes for this day
                            day_classes = classes_per_day
                            if i < remainder:
                                day_classes += 1
                            
                            # Calculate attended classes for this day (proportional)
                            day_attended = 0
                            if total > 0:
                                day_attended = round((attended * day_classes) / total)
                            
                            # Check if attendance record already exists for this student, subject, and date
                            existing_attendance = Attendance.query.filter_by(
                                student_id=student.id,
                                subject_id=subject_id,
                                date=date_str
                            ).first()
                            
                            if existing_attendance:
                                existing_attendance.total_classes = day_classes
                                existing_attendance.attended_classes = day_attended
                            else:
                                attendance = Attendance(
                                    student_id=student.id,
                                    subject_id=subject_id,
                                    date=date_str,
                                    week_number=0,  # Optionally calculate week number if needed
                                    total_classes=day_classes,
                                    attended_classes=day_attended
                                )
                                db.session.add(attendance)
        
        db.session.commit()
        flash('Attendance updated successfully!')
        return redirect(url_for('attendance.dashboard'))

    students_data = [
        {
            "id": s.id,
            "roll_no": s.roll_no,
            "reg_no": s.reg_no,
            "name": s.name
        }
        for s in students
    ]
    
    # Load existing attendance data for the selected week and subject
    existing_attendance = {}
    selected_subject_id = request.args.get('subject_id', type=int)
    selected_week_date = request.args.get('week_date')
    
    if selected_subject_id and selected_week_date:
        from datetime import datetime, timedelta
        try:
            selected_date = datetime.strptime(selected_week_date, '%Y-%m-%d')
            # Get Monday of the week
            days_since_monday = selected_date.weekday()
            monday = selected_date - timedelta(days=days_since_monday)
            
            # Get week dates (Monday to Friday, optionally Saturday)
            include_saturday = request.args.get('include_saturday') == 'true'
            week_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
            if include_saturday:
                week_days.append('Saturday')
            
            # Load existing attendance data for each student
            for student in students:
                total_classes = 0
                total_attended = 0
                
                # Sum up attendance for all days in the week
                for i, day in enumerate(week_days):
                    current_date = monday + timedelta(days=i)
                    date_str = current_date.strftime('%Y-%m-%d')
                    
                    # Get existing attendance record
                    attendance_record = Attendance.query.filter_by(
                        student_id=student.id,
                        subject_id=selected_subject_id,
                        date=date_str
                    ).first()
                    
                    if attendance_record:
                        total_classes += attendance_record.total_classes
                        total_attended += attendance_record.attended_classes
                
                existing_attendance[student.id] = {
                    'total_classes': total_classes,
                    'total_attended': total_attended,
                    'percentage': round((total_attended / total_classes) * 100, 2) if total_classes > 0 else 0
                }
        except ValueError:
            # Invalid date format, ignore
            pass
    
    return render_template('update_attendance.html', 
                         students=students_data, 
                         subjects=subjects, 
                         active_semester=active_semester,
                         existing_attendance=existing_attendance,
                         selected_subject_id=selected_subject_id,
                         selected_week_date=selected_week_date)

# Export full list with semester-wise breakdown
@attendance_bp.route('/export_all')
@login_required
def export_all():
    advisor = Advisor.query.get(session['advisor_id'])
    students = Student.query.filter_by(batch_id=advisor.batch_id).all()
    semesters = Semester.query.filter_by(batch_id=advisor.batch_id).all()

    data = []
    for student in students:
        student_row = {
            'Roll No': student.roll_no,
            'Registration No': student.reg_no,
            'Name': student.name
        }
        
        total_classes = 0
        total_attended = 0
        
        # Add semester-wise breakdown
        for semester in semesters:
            semester_subjects = Subject.query.filter_by(semester_id=semester.id).all()
            semester_total = 0
            semester_attended = 0
            
            for subject in semester_subjects:
                attendance_records = Attendance.query.filter_by(
                    student_id=student.id, 
                    subject_id=subject.id
                ).all()
                
                subject_total = sum(a.total_classes for a in attendance_records)
                subject_attended = sum(a.attended_classes for a in attendance_records)
                
                semester_total += subject_total
                semester_attended += subject_attended
            
            semester_percentage = round((semester_attended / semester_total) * 100, 2) if semester_total > 0 else 0
            
            student_row[f'{semester.name} - Total'] = semester_total
            student_row[f'{semester.name} - Attended'] = semester_attended
            student_row[f'{semester.name} - %'] = semester_percentage
            
            total_classes += semester_total
            total_attended += semester_attended
        
        # Add subject-wise breakdown
        for semester in semesters:
            semester_subjects = Subject.query.filter_by(semester_id=semester.id).all()
            for subject in semester_subjects:
                attendance_records = Attendance.query.filter_by(
                    student_id=student.id, 
                    subject_id=subject.id
                ).all()
                
                subject_total = sum(a.total_classes for a in attendance_records)
                subject_attended = sum(a.attended_classes for a in attendance_records)
                subject_percentage = round((subject_attended / subject_total) * 100, 2) if subject_total > 0 else 0
                
                # Create subject identifier
                subject_identifier = f'{semester.name} - {subject.name}'
                student_row[f'{subject_identifier} - Total'] = subject_total
                student_row[f'{subject_identifier} - Attended'] = subject_attended
                student_row[f'{subject_identifier} - %'] = subject_percentage
        
        overall_percentage = round((total_attended / total_classes) * 100, 2) if total_classes > 0 else 0
        
        student_row['Total Classes'] = total_classes
        student_row['Total Attended'] = total_attended
        student_row['Overall %'] = overall_percentage
        
        data.append(student_row)

    df = pd.DataFrame(data)
    
    # Create Excel writer with formatting
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Attendance Report')
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Attendance Report']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width (minimum 10, maximum 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Format percentage columns
        percentage_columns = [col for col in df.columns if '%' in col]
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in percentage_columns:
                for row_idx in range(2, len(df) + 2):  # Skip header row
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = '0.00%'
        
        # Enable text wrapping for all cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
    
    output.seek(0)
    return send_file(output, download_name='all_students_attendance.xlsx', as_attachment=True)

# Export ineligible list (<75%)
@attendance_bp.route('/export_ineligible')
@login_required
def export_ineligible():
    advisor = Advisor.query.get(session['advisor_id'])
    students = Student.query.filter_by(batch_id=advisor.batch_id).all()
    semesters = Semester.query.filter_by(batch_id=advisor.batch_id).all()
    
    # Get export parameters
    export_type = request.args.get('export_type', 'overall')  # 'overall' or 'subject'
    subject_id = request.args.get('subject_id', type=int)
    
    # Get all subjects for dropdown
    all_subjects = []
    for semester in semesters:
        semester_subjects = Subject.query.filter_by(semester_id=semester.id).all()
        for subject in semester_subjects:
            all_subjects.append({
                'id': subject.id,
                'name': f"{semester.name} - {subject.name}",
                'semester': semester.name,
                'subject_name': subject.name
            })

    data = []
    for student in students:
        total_classes = 0
        total_attended = 0
        
        if export_type == 'overall':
            # Overall attendance calculation
            for semester in semesters:
                semester_subjects = Subject.query.filter_by(semester_id=semester.id).all()
                for subject in semester_subjects:
                    attendance_records = Attendance.query.filter_by(
                        student_id=student.id, 
                        subject_id=subject.id
                    ).all()
                    
                    total_classes += sum(a.total_classes for a in attendance_records)
                    total_attended += sum(a.attended_classes for a in attendance_records)
            
            overall_percentage = round((total_attended / total_classes) * 100, 2) if total_classes > 0 else 0

            if overall_percentage < 75:
                student_row = {
                    'Roll No': student.roll_no,
                    'Registration No': student.reg_no,
                    'Name': student.name,
                    'Total Classes': total_classes,
                    'Total Attended': total_attended,
                    'Overall %': overall_percentage
                }
                
                # Add semester-wise breakdown
                for semester in semesters:
                    semester_subjects = Subject.query.filter_by(semester_id=semester.id).all()
                    semester_total = 0
                    semester_attended = 0
                    
                    for subject in semester_subjects:
                        attendance_records = Attendance.query.filter_by(
                            student_id=student.id, 
                            subject_id=subject.id
                        ).all()
                        
                        semester_total += sum(a.total_classes for a in attendance_records)
                        semester_attended += sum(a.attended_classes for a in attendance_records)
                    
                    semester_percentage = round((semester_attended / semester_total) * 100, 2) if semester_total > 0 else 0
                    
                    student_row[f'{semester.name} - Total'] = semester_total
                    student_row[f'{semester.name} - Attended'] = semester_attended
                    student_row[f'{semester.name} - %'] = semester_percentage
                
                # Add subject-wise breakdown for ineligible students
                for semester in semesters:
                    semester_subjects = Subject.query.filter_by(semester_id=semester.id).all()
                    for subject in semester_subjects:
                        attendance_records = Attendance.query.filter_by(
                            student_id=student.id, 
                            subject_id=subject.id
                        ).all()
                        
                        subject_total = sum(a.total_classes for a in attendance_records)
                        subject_attended = sum(a.attended_classes for a in attendance_records)
                        subject_percentage = round((subject_attended / subject_total) * 100, 2) if subject_total > 0 else 0
                        
                        # Create subject identifier
                        subject_identifier = f'{semester.name} - {subject.name}'
                        student_row[f'{subject_identifier} - Total'] = subject_total
                        student_row[f'{subject_identifier} - Attended'] = subject_attended
                        student_row[f'{subject_identifier} - %'] = subject_percentage
                
                data.append(student_row)
        
        elif export_type == 'subject' and subject_id:
            # Subject-specific attendance calculation
            subject = Subject.query.get(subject_id)
            if subject:
                attendance_records = Attendance.query.filter_by(
                    student_id=student.id, 
                    subject_id=subject_id
                ).all()
                
                subject_total = sum(a.total_classes for a in attendance_records)
                subject_attended = sum(a.attended_classes for a in attendance_records)
                subject_percentage = round((subject_attended / subject_total) * 100, 2) if subject_total > 0 else 0
                
                if subject_percentage < 75:
                    student_row = {
                        'Roll No': student.roll_no,
                        'Registration No': student.reg_no,
                        'Name': student.name,
                        'Subject': f"{subject.semester.name} - {subject.name}",
                        'Total Classes': subject_total,
                        'Attended Classes': subject_attended,
                        'Attendance %': subject_percentage
                    }
                    data.append(student_row)

    df = pd.DataFrame(data)
    
    # Create Excel writer with formatting
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = 'Ineligible Students - Overall' if export_type == 'overall' else f'Ineligible Students - {subject.name if subject_id else "Subject"}'
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width (minimum 10, maximum 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")  # Red for ineligible
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Format percentage columns
        percentage_columns = [col for col in df.columns if '%' in col]
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in percentage_columns:
                for row_idx in range(2, len(df) + 2):  # Skip header row
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = '0.00%'
        
        # Enable text wrapping for all cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
    
    output.seek(0)
    
    # Return template with export options if no export type specified
    if not request.args.get('export_type'):
        return render_template('export_ineligible.html', subjects=all_subjects)
    
    filename = 'ineligible_students_overall.xlsx' if export_type == 'overall' else f'ineligible_students_{subject.name.replace(" ", "_")}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True)

@attendance_bp.route('/weeks_summary')
@login_required
def weeks_summary():
    advisor = Advisor.query.get(session['advisor_id'])
    active_semester = Semester.query.filter_by(batch_id=advisor.batch_id, is_active=True).first()
    weeks_list = []
    weeks_count = 0
    if active_semester:
        semester_subjects = Subject.query.filter_by(semester_id=active_semester.id).all()
        subject_ids = [s.id for s in semester_subjects]
        if subject_ids:
            from datetime import timedelta
            week_dates = db.session.query(Attendance.date).filter(
                Attendance.subject_id.in_(subject_ids)
            ).distinct().all()
            mondays = set()
            for (date_obj,) in week_dates:
                if date_obj:
                    monday = date_obj - timedelta(days=date_obj.weekday())
                    mondays.add(monday)
            weeks_list = sorted(list(mondays))
            weeks_count = len(weeks_list)
    return render_template('weeks_summary.html', weeks_list=weeks_list, weeks_count=weeks_count, active_semester=active_semester)
